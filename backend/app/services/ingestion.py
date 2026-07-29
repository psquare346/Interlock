"""Catalog ingestion: parse, column mapping, UoM normalization, row
validation, price-change surveillance.

Implements the pipeline in files/catalog-ingestion-agent.md. Layer 1 (rules
and lookup tables) is fully implemented and is the only layer that runs in
local mode. The smart-model column mapper (services/llm.py) is consulted only
when rules leave required fields unmapped AND a key is configured. Missing AI
degrades coverage, never correctness: unresolved rows queue for a human.

Hard validation failures can never be overridden by a confidence score.
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation

from sqlalchemy import select
from sqlalchemy.orm import Session

from ..config import get_settings
from ..models import (
    CatalogItem, CatalogVersion, ItemState, Supplier, VersionStatus,
)
from . import llm

# --------------------------------------------------------------------------
# Canonical schema
# --------------------------------------------------------------------------

REQUIRED_FIELDS = {"supplier_part_id", "description", "uom", "unit_price", "currency"}

CANONICAL_FIELDS = REQUIRED_FIELDS | {
    "price_unit", "unspsc", "material_group", "manufacturer",
    "manufacturer_part_id", "lead_time_days", "long_description",
}

# Header synonyms, keyed by normalized form (lowercase, umlauts folded,
# non-alphanumerics stripped). English + German, since EU supplier files are
# the second-most-common style in the golden set.
_SYNONYMS: dict[str, tuple[str, float]] = {}


def _learn(target: str, confidence: float, *headers: str) -> None:
    for h in headers:
        _SYNONYMS[_norm(h)] = (target, confidence)


def _norm(header: str) -> str:
    h = header.strip().lower()
    for a, b in (("ä", "a"), ("ö", "o"), ("ü", "u"), ("ß", "ss")):
        h = h.replace(a, b)
    return re.sub(r"[^a-z0-9]", "", h)


_learn("supplier_part_id", 0.98,
       "supplier part id", "part number", "part no", "part#", "partno", "sku",
       "item number", "item no", "artikelnummer", "art.-nr.", "art-nr", "artnr",
       "bestellnummer", "lieferanten-artikelnummer", "item id", "product code")
_learn("description", 0.97,
       "description", "item description", "product name", "bezeichnung",
       "artikelbezeichnung", "kurztext", "beschreibung", "name", "desc")
_learn("long_description", 0.95, "long description", "langtext", "detail description")
_learn("uom", 0.97,
       "uom", "unit of measure", "unit", "me", "mengeneinheit", "einheit",
       "order unit", "bestelleinheit", "uom code")
_learn("unit_price", 0.97,
       "unit price", "price", "preis", "einzelpreis", "nettopreis", "unit cost",
       "list price", "vk-preis", "price usd", "price eur")
_learn("price_unit", 0.95,
       "price unit", "vpe", "preiseinheit", "verpackungseinheit", "per",
       "price per", "pricing qty")
_learn("currency", 0.98, "currency", "wahrung", "waehrung", "curr", "ccy")
_learn("unspsc", 0.97, "unspsc", "unspsc code", "commodity code")
_learn("material_group", 0.95, "material group", "matgroup", "warengruppe", "matkl")
_learn("manufacturer", 0.96, "manufacturer", "hersteller", "mfr", "brand", "marke")
_learn("manufacturer_part_id", 0.96,
       "manufacturer part number", "mpn", "mfr part no", "herstellernummer",
       "hersteller-artikelnummer", "manufacturer part")
_learn("lead_time_days", 0.94, "lead time", "lead time days", "lieferzeit",
       "delivery time", "delivery days")

# --------------------------------------------------------------------------
# UoM lookup — the local mirror of SAP T006 (Layer 1 of the cascade)
# --------------------------------------------------------------------------

SAP_UOMS = {
    "EA", "PAK", "KAR", "KG", "G", "L", "ML", "M", "MM", "CM", "ROL",
    "PAA", "DZ", "SET", "BT", "TU", "KI",
}

_UOM_MAP = {
    # each
    "ea": "EA", "each": "EA", "pc": "EA", "pcs": "EA", "pce": "EA", "st": "EA",
    "stk": "EA", "stuck": "EA", "stck": "EA", "piece": "EA", "pieces": "EA",
    "unit": "EA", "un": "EA", "c62": "EA",
    # pack / carton / box
    "pk": "PAK", "pak": "PAK", "pack": "PAK", "pkg": "PAK", "pck": "PAK",
    "bx": "KAR", "box": "KAR", "kar": "KAR", "karton": "KAR", "ct": "KAR",
    "ctn": "KAR", "cs": "KAR", "case": "KAR", "ki": "KI", "kiste": "KI",
    # weight / volume / length
    "kg": "KG", "kilogram": "KG", "kilo": "KG", "g": "G", "gram": "G",
    "l": "L", "liter": "L", "litre": "L", "ltr": "L", "ml": "ML",
    "m": "M", "meter": "M", "metre": "M", "mtr": "M", "mm": "MM", "cm": "CM",
    # misc
    "rol": "ROL", "roll": "ROL", "rolle": "ROL", "rl": "ROL",
    "pr": "PAA", "pair": "PAA", "paar": "PAA", "paa": "PAA",
    "dz": "DZ", "doz": "DZ", "dozen": "DZ", "dtz": "DZ",
    "set": "SET", "satz": "SET", "kit": "SET",
    "bt": "BT", "bottle": "BT", "fl": "BT", "flasche": "BT", "tu": "TU", "tube": "TU",
}

_ISO_4217 = {
    "USD", "EUR", "GBP", "CHF", "JPY", "CNY", "CAD", "MXN", "AUD", "SEK",
    "NOK", "DKK", "PLN", "CZK", "HUF", "INR", "SGD", "HKD", "KRW", "BRL",
}


def lookup_uom(raw: str | None) -> str | None:
    """Layer 1: lookup table -> SAP unit. ~95% of rows end here."""
    if not raw:
        return None
    key = re.sub(r"[^a-z0-9]", "", raw.strip().lower())
    if raw.strip().upper() in SAP_UOMS:
        return raw.strip().upper()
    return _UOM_MAP.get(key)


# --------------------------------------------------------------------------
# Step 1 — Parse (code, no model)
# --------------------------------------------------------------------------

class IngestError(Exception):
    pass


def parse_file(filename: str, content: bytes) -> tuple[list[str], list[dict]]:
    """Detect encoding and delimiter, return (headers, rows-as-dicts).

    CSV and XLSX. Ariba CIF and BMEcat are on the roadmap (agent spec input
    list); they land here as new branches with the same return shape.
    """
    if filename.lower().endswith((".xlsx", ".xlsm")):
        return _parse_xlsx(content)
    return _parse_csv(content)


def _parse_csv(content: bytes) -> tuple[list[str], list[dict]]:
    text = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise IngestError("Could not decode file as UTF-8 or Windows-1252")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        # Sniffer gives up on single-column files; pick the most frequent.
        delimiter = max(",;\t|", key=sample.count)

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    raw_rows = [r for r in reader if any(c.strip() for c in r)]
    if not raw_rows:
        raise IngestError("File contains no data rows")

    headers = [h.strip() for h in raw_rows[0]]
    rows = []
    for raw in raw_rows[1:]:
        row = {headers[i]: (raw[i].strip() if i < len(raw) else "") for i in range(len(headers))}
        rows.append(row)
    return headers, rows


def _parse_xlsx(content: bytes) -> tuple[list[str], list[dict]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    iterator = ws.iter_rows(values_only=True)
    headers = None
    rows: list[dict] = []
    for values in iterator:
        if headers is None:
            if values and any(v not in (None, "") for v in values):
                headers = [str(v).strip() if v is not None else "" for v in values]
            continue
        if not values or all(v in (None, "") for v in values):
            continue
        rows.append({
            headers[i]: ("" if i >= len(values) or values[i] is None else str(values[i]).strip())
            for i in range(len(headers))
        })
    if headers is None:
        raise IngestError("Workbook has no header row")
    return headers, rows


# --------------------------------------------------------------------------
# Step 2 — Column mapping
# --------------------------------------------------------------------------

@dataclass
class MappingResult:
    mappings: list[dict] = field(default_factory=list)   # source_column/target_field/confidence/rationale
    unmapped_columns: list[str] = field(default_factory=list)
    missing_required: list[str] = field(default_factory=list)
    source: str = "rules"   # rules | llm | confirmed

    def as_dict(self) -> dict:
        return {
            "mappings": self.mappings,
            "unmapped_columns": self.unmapped_columns,
            "missing_required": self.missing_required,
            "source": self.source,
        }

    def column_for(self, target: str) -> str | None:
        for m in self.mappings:
            if m["target_field"] == target:
                return m["source_column"]
        return None


def map_columns(headers: list[str], sample_rows: list[dict],
                confirmed: dict | None = None) -> MappingResult:
    """Rules first; a confirmed supplier mapping skips everything; the smart
    model is consulted only for required fields the rules could not place."""
    if confirmed:
        result = MappingResult(source="confirmed")
        result.mappings = confirmed.get("mappings", [])
        mapped_targets = {m["target_field"] for m in result.mappings}
        result.missing_required = sorted(REQUIRED_FIELDS - mapped_targets)
        mapped_sources = {m["source_column"] for m in result.mappings}
        result.unmapped_columns = [h for h in headers if h not in mapped_sources]
        return result

    result = MappingResult()
    taken: set[str] = set()
    for h in headers:
        hit = _SYNONYMS.get(_norm(h))
        if hit and hit[0] not in taken:
            target, conf = hit
            result.mappings.append({
                "source_column": h, "target_field": target,
                "confidence": conf, "rationale": "synonym table",
            })
            taken.add(target)
        else:
            result.unmapped_columns.append(h)

    result.missing_required = sorted(REQUIRED_FIELDS - taken)

    # Escalate to the smart model only when rules failed on something required.
    if result.missing_required:
        llm_result = llm.map_columns(headers, sample_rows, sorted(CANONICAL_FIELDS))
        if llm_result:
            for m in llm_result.get("mappings", []):
                if m["target_field"] in taken or m["source_column"] not in result.unmapped_columns:
                    continue
                result.mappings.append(m)
                taken.add(m["target_field"])
                result.unmapped_columns.remove(m["source_column"])
            result.missing_required = sorted(REQUIRED_FIELDS - taken)
            result.source = "llm"

    return result


# --------------------------------------------------------------------------
# Steps 3-5 — Normalize, validate, surveil
# --------------------------------------------------------------------------

_NUM_CLEAN = re.compile(r"[^\d,.\-]")


def _parse_decimal(raw: str) -> Decimal | None:
    """Handle 1,234.56 and 1.234,56 and 42,00."""
    s = _NUM_CLEAN.sub("", raw or "").strip()
    if not s:
        return None
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):     # 1.234,56 — German
            s = s.replace(".", "").replace(",", ".")
        else:                                # 1,234.56 — English
            s = s.replace(",", "")
    elif "," in s:
        # A lone comma followed by exactly 2 digits is a decimal comma.
        head, _, tail = s.rpartition(",")
        s = f"{head.replace(',', '')}.{tail}" if len(tail) == 2 else s.replace(",", "")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def ingest(
    db: Session,
    tenant_id: str,
    supplier: Supplier,
    filename: str,
    content: bytes,
) -> CatalogVersion:
    """Run the full pipeline and persist a CatalogVersion with per-row states."""
    settings = get_settings()
    headers, rows = parse_file(filename, content)
    if not rows:
        raise IngestError("No data rows after the header")

    mapping = map_columns(headers, rows[:20], supplier.confirmed_mapping)

    prior_no = db.scalar(
        select(CatalogVersion.version_no)
        .where(CatalogVersion.supplier_id == supplier.id)
        .order_by(CatalogVersion.version_no.desc())
        .limit(1)
    ) or 0

    version = CatalogVersion(
        tenant_id=tenant_id,
        supplier_id=supplier.id,
        version_no=prior_no + 1,
        source_filename=filename,
        mapping_used=mapping.as_dict(),
        row_count=len(rows),
    )
    db.add(version)
    db.flush()

    # Previous published prices for surveillance (step 5).
    prev_prices = _published_prices(db, supplier.id)

    col = {t: mapping.column_for(t) for t in CANONICAL_FIELDS}
    mapping_conf = {
        m["target_field"]: float(m.get("confidence", 0.9)) for m in mapping.mappings
    }
    seen_parts: set[str] = set()

    for raw in rows:
        item, confidence, reasons, hard_fails = _normalize_row(
            raw, col, mapping_conf, mapping.missing_required
        )
        item.tenant_id = tenant_id
        item.version_id = version.id
        item.supplier_id = supplier.id

        # Uniqueness within the version is a hard rule.
        if item.supplier_part_id:
            if item.supplier_part_id in seen_parts:
                hard_fails.append(f"Duplicate supplier_part_id {item.supplier_part_id!r} in file")
            seen_parts.add(item.supplier_part_id)

        item.price_flags = _surveil(item, prev_prices, settings.PRICE_CHANGE_FLAG_PCT)
        if item.price_flags:
            reasons.extend(item.price_flags)
            confidence = min(confidence, settings.AUTO_APPROVE_THRESHOLD - 0.01)

        if hard_fails:
            item.state = ItemState.MANUAL
            reasons = hard_fails + reasons
            confidence = 0.0
        elif confidence >= settings.AUTO_APPROVE_THRESHOLD:
            item.state = ItemState.AUTO_APPROVED
        elif confidence >= settings.REVIEW_THRESHOLD:
            item.state = ItemState.NEEDS_REVIEW
        else:
            item.state = ItemState.MANUAL

        item.confidence = round(confidence, 3)
        item.review_reasons = reasons or None
        db.add(item)

    db.commit()
    db.refresh(version)
    return version


def _normalize_row(
    raw: dict,
    col: dict[str, str | None],
    mapping_conf: dict[str, float],
    missing_required: list[str],
) -> tuple[CatalogItem, float, list[str], list[str]]:
    def get(target: str) -> str:
        c = col.get(target)
        return (raw.get(c) or "").strip() if c else ""

    reasons: list[str] = []
    hard: list[str] = [f"Required field never mapped: {m}" for m in missing_required]
    confidences: list[float] = [mapping_conf.get(t, 0.9) for t in mapping_conf]

    item = CatalogItem(raw_row=raw)

    # supplier_part_id — never invented; absence is manual, always.
    item.supplier_part_id = get("supplier_part_id") or None
    if not item.supplier_part_id:
        hard.append("Missing supplier_part_id")

    # description — truncate to 40 (SAP TXZ01), spill to long text.
    desc = get("description")
    if not desc:
        hard.append("Missing description")
    item.description = desc[:40] or None
    long_desc = get("long_description")
    if len(desc) > 40:
        long_desc = long_desc or desc
        reasons.append("Description over 40 chars; spilled to long text")
    item.long_description = long_desc or None

    # UoM cascade, Layer 1 only in local mode.
    item.uom_raw = get("uom") or None
    item.uom_sap = lookup_uom(item.uom_raw)
    if item.uom_raw and not item.uom_sap:
        hard.append(f"Unit {item.uom_raw!r} not in the T006 mirror")
    elif not item.uom_raw:
        hard.append("Missing unit of measure")
    elif item.uom_raw.strip().upper() != item.uom_sap:
        confidences.append(0.93)
        reasons.append(f"UoM normalized {item.uom_raw!r} -> {item.uom_sap}")

    # Price — parsed by code and validated, full stop.
    price = _parse_decimal(get("unit_price"))
    if price is None or price <= 0:
        hard.append(f"unit_price does not parse as a positive decimal: {get('unit_price')!r}")
    else:
        item.unit_price = price

    pu_raw = get("price_unit")
    if pu_raw:
        pu = _parse_decimal(pu_raw)
        if pu is None or pu <= 0 or pu != pu.to_integral_value():
            hard.append(f"price_unit is not a positive integer: {pu_raw!r}")
        else:
            item.price_unit = int(pu)
    else:
        item.price_unit = 1

    currency = get("currency").upper()
    if currency in _ISO_4217:
        item.currency = currency
    else:
        hard.append(f"Currency {currency!r} is not valid ISO 4217")

    # Optional fields.
    unspsc = re.sub(r"\D", "", get("unspsc"))
    if unspsc and len(unspsc) in (8, 10):
        item.unspsc = unspsc
        confidences.append(0.97)
    elif get("unspsc"):
        reasons.append(f"Supplied UNSPSC {get('unspsc')!r} is malformed")
        confidences.append(0.75)
    else:
        # No classifier without an AI key: queue instead of guessing.
        reasons.append("No UNSPSC supplied; classifier unavailable in rule-only mode")
        confidences.append(0.85)

    # Material group is never auto-approved — it drives GL account determination.
    mg = get("material_group")
    if mg:
        item.material_group = mg
        reasons.append("Material group supplied by file; verify against master")
        confidences.append(0.90)

    item.manufacturer = get("manufacturer") or None
    item.manufacturer_part_id = get("manufacturer_part_id") or None
    lead = _parse_decimal(get("lead_time_days"))
    item.lead_time_days = int(lead) if lead is not None and lead >= 0 else None

    confidence = min(confidences) if confidences else 0.5
    return item, confidence, reasons, hard


def revalidate_item(db: Session, item: CatalogItem) -> list[str]:
    """Re-run the hard rules after a human edit. Same rules as ingestion —
    an edit can fix a row, it can never bypass validation."""
    hard: list[str] = []

    if not item.supplier_part_id:
        hard.append("Missing supplier_part_id")
    else:
        twin = db.scalars(
            select(CatalogItem).where(
                CatalogItem.version_id == item.version_id,
                CatalogItem.supplier_part_id == item.supplier_part_id,
                CatalogItem.id != item.id,
            )
        ).first()
        if twin is not None:
            hard.append(f"Duplicate supplier_part_id {item.supplier_part_id!r} in version")

    if not item.description:
        hard.append("Missing description")

    if item.uom_raw:
        item.uom_sap = lookup_uom(item.uom_raw)
        if not item.uom_sap:
            hard.append(f"Unit {item.uom_raw!r} not in the T006 mirror")
    else:
        hard.append("Missing unit of measure")

    if item.unit_price is None or Decimal(str(item.unit_price)) <= 0:
        hard.append("unit_price must be a positive decimal")
    if not item.price_unit or item.price_unit < 1:
        hard.append("price_unit must be a positive integer")
    if (item.currency or "") not in _ISO_4217:
        hard.append(f"Currency {item.currency!r} is not valid ISO 4217")

    return hard


def _published_prices(db: Session, supplier_id: str) -> dict[str, Decimal]:
    version = db.scalars(
        select(CatalogVersion)
        .where(
            CatalogVersion.supplier_id == supplier_id,
            CatalogVersion.status == VersionStatus.PUBLISHED,
        )
        .order_by(CatalogVersion.version_no.desc())
    ).first()
    if version is None:
        return {}
    return {
        i.supplier_part_id: Decimal(str(i.unit_price))
        for i in version.items
        if i.supplier_part_id and i.unit_price is not None
        and i.state in (ItemState.AUTO_APPROVED, ItemState.APPROVED)
    }


def _surveil(item: CatalogItem, prev: dict[str, Decimal], flag_pct: float) -> list[str]:
    """Step 5 — statistical, not a model."""
    if not item.supplier_part_id or item.supplier_part_id not in prev:
        return []
    old = prev[item.supplier_part_id]
    new = Decimal(str(item.unit_price)) if item.unit_price is not None else None
    flags: list[str] = []
    if new is None or new == 0:
        if old != 0:
            flags.append(f"Price moved to zero/unparseable from {old}")
        return flags
    if old == 0:
        flags.append(f"Price moved from zero to {new}")
        return flags
    delta_pct = (new - old) / old * 100
    if delta_pct > Decimal(str(flag_pct)):
        flags.append(f"Price increase {delta_pct.quantize(Decimal('0.1'))}% ({old} -> {new})")
    return flags
